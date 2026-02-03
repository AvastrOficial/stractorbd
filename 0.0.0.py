import json
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog
import os
import sys
import time
from openpyxl import load_workbook
import datetime

def mostrar_progreso_consola(etapa, porcentaje, detalles=""):
    """Muestra progreso en la consola"""
    barra = '█' * int(porcentaje/5) + '░' * (20 - int(porcentaje/5))
    print(f"\r{etapa:25} [{barra}] {porcentaje:6.1f}% | {detalles}", end="")
    sys.stdout.flush()
    
    if porcentaje >= 100:
        print()

def columna_a_letra(col):
    """Convierte número de columna a letra de Excel (A, B, C, ..., Z, AA, AB, etc.)"""
    letra = ''
    while col > 0:
        col, resto = divmod(col - 1, 26)
        letra = chr(65 + resto) + letra
    return letra

def preguntar_numero_filas_columnas(total_filas, total_columnas):
    """Pregunta al usuario cuántas filas y columnas quiere procesar"""
    root = tk.Tk()
    root.withdraw()
    
    config_window = tk.Toplevel(root)
    config_window.title("Configurar Muestra")
    config_window.geometry("500x350")
    
    # Centrar
    config_window.update_idletasks()
    x = (config_window.winfo_screenwidth() // 2) - 250
    y = (config_window.winfo_screenheight() // 2) - 175
    config_window.geometry(f'500x350+{x}+{y}')
    
    tk.Label(config_window, text="⚙️ CONFIGURAR MUESTRA RÁPIDA", 
             font=("Arial", 14, "bold"), pady=10).pack()
    
    tk.Label(config_window, 
             text=f"Archivo tiene: {total_filas} filas, {total_columnas} columnas",
             font=("Arial", 10)).pack(pady=5)
    
    # Variables para almacenar resultado
    max_filas = min(10000, total_filas)  # Cambiado a 10,000
    resultado = {"filas": min(1000, total_filas), "columnas": min(22, total_columnas)}
    
    # Frame para filas
    frame_filas = tk.Frame(config_window, pady=15)
    frame_filas.pack()
    
    tk.Label(frame_filas, text="Filas a procesar:", 
             font=("Arial", 10, "bold")).pack(side='left', padx=5)
    
    filas_var = tk.StringVar(value="10000")
    entry_filas = tk.Entry(frame_filas, textvariable=filas_var, 
                          width=15, font=("Arial", 10))
    entry_filas.pack(side='left', padx=5)
    
    tk.Label(frame_filas, 
             text=f"(Máximo: {max_filas})",
             font=("Arial", 9), fg="blue").pack(side='left', padx=5)
    
    # Lista de etiquetas requeridas
    etiquetas_frame = tk.Frame(config_window, relief="groove", borderwidth=2, padx=10, pady=10)
    etiquetas_frame.pack(pady=10, padx=20, fill="x")
    
    tk.Label(etiquetas_frame, text="📋 ETIQUETAS REQUERIDAS:", 
             font=("Arial", 10, "bold")).pack(anchor='w')
    
    etiquetas = ["CVE", "NOMBRE", "PATERNO", "MATERNO", "FECNAC", "SEXO", 
                "CALLE", "INT", "EXT", "COLONIA", "CP", "E", "D", "M", 
                "S", "L", "MZA", "CONSEC", "CRED", "FOLIO", "NAC", "CURP"]
    
    # Mostrar etiquetas en dos columnas
    frame_columnas = tk.Frame(etiquetas_frame)
    frame_columnas.pack(pady=5)
    
    mitad = len(etiquetas) // 2
    for i in range(mitad):
        tk.Label(frame_columnas, text=f"• {etiquetas[i]:15}", 
                font=("Courier", 9)).grid(row=i, column=0, sticky='w', padx=5)
    for i in range(mitad, len(etiquetas)):
        tk.Label(frame_columnas, text=f"• {etiquetas[i]:15}", 
                font=("Courier", 9)).grid(row=i-mitad, column=1, sticky='w', padx=5)
    
    tk.Label(etiquetas_frame, text="Total: 22 columnas", 
             font=("Arial", 9), fg="green").pack(anchor='w', pady=5)
    
    # Información de configuración fija para columnas
    info_frame = tk.Frame(config_window, pady=10)
    info_frame.pack()
    
    tk.Label(info_frame, text="⚠️  NOTA: Columnas fijas en 22 (etiquetas requeridas)",
             font=("Arial", 9), fg="red").pack()
    
    def aceptar():
        try:
            filas = int(filas_var.get())
            
            # Validar filas (1-10,000)
            if 1 <= filas <= max_filas:
                resultado["filas"] = filas
                resultado["columnas"] = 22  # Fijo para las etiquetas
                config_window.destroy()
            else:
                messagebox.showerror("Error", 
                    f"Valores fuera de rango.\n"
                    f"Filas: 1-{max_filas}\n"
                    f"Columnas: 22 (fijo por etiquetas requeridas)")
        except ValueError:
            messagebox.showerror("Error", "Por favor ingresa un número válido para filas")
    
    def cancelar():
        resultado["filas"] = None
        resultado["columnas"] = None
        config_window.destroy()
    
    # Botones
    frame_botones = tk.Frame(config_window, pady=20)
    frame_botones.pack()
    
    tk.Button(frame_botones, text="✅ Procesar (10,000 filas)", command=aceptar,
              width=25, bg="#4CAF50", fg="white", font=("Arial", 10, "bold")).pack(side='left', padx=10)
    tk.Button(frame_botones, text="❌ Cancelar", command=cancelar,
              width=15).pack(side='left', padx=10)
    
    # Hacer focus en el campo de filas
    entry_filas.focus_set()
    entry_filas.select_range(0, tk.END)
    
    # Esperar a que se cierre la ventana
    config_window.wait_window(config_window)
    
    return resultado["filas"], resultado["columnas"]

def seleccionar_destino_guardado(nombre_predeterminado):
    """Permite al usuario seleccionar dónde guardar el archivo"""
    root = tk.Tk()
    root.withdraw()
    
    archivo = filedialog.asksaveasfilename(
        title="Guardar archivo JSON",
        initialfile=nombre_predeterminado,
        defaultextension=".json",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
    )
    
    return archivo

def procesar_excel_completo(archivo_excel, destino_guardado=None):
    """Procesa el archivo Excel completo"""
    try:
        print("\n" + "="*60)
        print("PROCESANDO ARCHIVO EXCEL - VERSIÓN COMPLETA")
        print("="*60)
        
        # Etapa 1: Cargar archivo
        mostrar_progreso_consola("Cargando archivo", 10, "Leyendo Excel...")
        wb = load_workbook(archivo_excel, data_only=True, read_only=True)
        
        # Etapa 2: Analizar estructura
        mostrar_progreso_consola("Analizando estructura", 30, f"{len(wb.sheetnames)} hojas")
        
        datos = {
            "archivo": os.path.basename(archivo_excel),
            "fecha_procesamiento": time.strftime("%Y-%m-%d %H:%M:%S"),
            "hojas": []
        }
        
        # Procesar cada hoja
        total_hojas = len(wb.sheetnames)
        for idx, sheet_name in enumerate(wb.sheetnames):
            progreso = 30 + (idx / total_hojas * 60)
            mostrar_progreso_consola(f"Procesando hoja {idx+1}", progreso, sheet_name)
            
            ws = wb[sheet_name]
            hoja_data = {
                "nombre": sheet_name,
                "filas_totales": ws.max_row,
                "columnas_totales": ws.max_column,
                "celdas_procesadas": 0,
                "datos": []
            }
            
            # Procesar TODAS las filas y columnas
            for fila in range(1, ws.max_row + 1):
                fila_datos = []
                for col in range(1, ws.max_column + 1):
                    valor = ws.cell(row=fila, column=col).value
                    if valor is not None:
                        valor_str = str(valor) if valor is not None else ""
                        fila_datos.append({
                            "columna": col,
                            "letra_col": columna_a_letra(col),
                            "valor": valor_str
                        })
                        hoja_data["celdas_procesadas"] += 1
                
                if fila_datos:
                    hoja_data["datos"].append({
                        "fila": fila,
                        "celdas": fila_datos
                    })
            
            datos["hojas"].append(hoja_data)
        
        wb.close()
        
        # Etapa 3: Guardar JSON
        mostrar_progreso_consola("Guardando JSON", 95, "Creando archivo...")
        
        # Generar nombre por defecto
        nombre_base = os.path.splitext(os.path.basename(archivo_excel))[0]
        nombre_predeterminado = f"{nombre_base}_COMPLETO_{time.strftime('%Y%m%d_%H%M%S')}.json"
        
        # Si no se proporcionó destino, preguntar al usuario
        if not destino_guardado:
            destino_guardado = seleccionar_destino_guardado(nombre_predeterminado)
            
            if not destino_guardado:  # Usuario canceló
                print("\n❌ Guardado cancelado por el usuario.")
                return None, datos
        
        # Asegurar extensión .json
        if not destino_guardado.lower().endswith('.json'):
            destino_guardado += '.json'
        
        with open(destino_guardado, 'w', encoding='utf-8') as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        
        total_celdas = sum(h["celdas_procesadas"] for h in datos["hojas"])
        total_filas = sum(len(h["datos"]) for h in datos["hojas"])
        mostrar_progreso_consola("✅ COMPLETADO", 100, 
                               f"{total_filas} filas, {total_celdas} celdas")
        
        return destino_guardado, datos
        
    except Exception as e:
        raise Exception(f"Error: {str(e)}")

def procesar_muestra_rapida(archivo_excel, filas=10000, columnas=22, destino_guardado=None):
    """Procesa solo una muestra del archivo con etiquetas específicas"""
    try:
        print("\n" + "="*60)
        print(f"PROCESANDO MUESTRA RÁPIDA ({filas} filas × {columnas} columnas)")
        print("="*60)
        
        mostrar_progreso_consola("Cargando archivo", 10, "Abriendo Excel...")
        wb = load_workbook(archivo_excel, data_only=True, read_only=True)
        ws = wb.active  # Solo primera hoja
        
        print(f"📊 Información de la hoja: {ws.title}")
        print(f"   - Filas totales: {ws.max_row}")
        print(f"   - Columnas totales: {ws.max_column}")
        print(f"   - Filas a procesar: {min(filas, ws.max_row)}")
        print(f"   - Columnas a procesar: {min(columnas, ws.max_column)}")
        
        # Definir las etiquetas específicas
        etiquetas = ["CVE", "NOMBRE", "PATERNO", "MATERNO", "FECNAC", "SEXO", 
                    "CALLE", "INT", "EXT", "COLONIA", "CP", "E", "D", "M", 
                    "S", "L", "MZA", "CONSEC", "CRED", "FOLIO", "NAC", "CURP"]
        
        print(f"📋 Etiquetas configuradas: {len(etiquetas)} columnas")
        
        muestra = {
            "archivo": os.path.basename(archivo_excel),
            "fecha_procesamiento": time.strftime("%Y-%m-%d %H:%M:%S"),
            "hoja": ws.title,
            "total_filas": ws.max_row,
            "total_columnas": ws.max_column,
            "etiquetas": etiquetas,
            "configuracion": {
                "filas_solicitadas": filas,
                "columnas_solicitadas": columnas,
                "filas_procesadas": min(filas, ws.max_row),
                "columnas_procesadas": min(columnas, ws.max_column)
            },
            "datos": []
        }
        
        # Procesar muestra - RESPETANDO los valores configurados
        filas_a_procesar = min(filas, ws.max_row)
        columnas_a_procesar = min(columnas, ws.max_column)
        
        print(f"\n📈 Procesando {filas_a_procesar} filas y {columnas_a_procesar} columnas...")
        print(f"🏷️  Etiquetas: {', '.join(etiquetas[:5])}...")
        
        for fila_idx, fila in enumerate(range(1, filas_a_procesar + 1)):
            progreso = 20 + (fila_idx / filas_a_procesar * 70)
            if fila_idx % 100 == 0 or fila_idx == filas_a_procesar - 1:
                mostrar_progreso_consola(f"Procesando fila {fila}", progreso, 
                                       f"{columnas_a_procesar} etiquetas")
            
            fila_datos = {}
            # Procesar las primeras 'columnas' columnas con etiquetas específicas
            for col in range(1, columnas_a_procesar + 1):
                if col-1 < len(etiquetas):
                    etiqueta = etiquetas[col-1]
                else:
                    etiqueta = f"COL_{col}"
                
                valor = ws.cell(row=fila, column=col).value
                
                if valor is not None:
                    # Convertir diferentes tipos de datos
                    if isinstance(valor, (int, float)):
                        fila_datos[etiqueta] = valor
                    elif isinstance(valor, (datetime.datetime, datetime.date)):
                        fila_datos[etiqueta] = valor.strftime("%Y-%m-%d")
                    else:
                        fila_datos[etiqueta] = str(valor).strip()
                else:
                    # Mantener celdas vacías como null
                    fila_datos[etiqueta] = None
            
            # Siempre agregar la fila, incluso si está vacía
            muestra["datos"].append({
                "fila": fila,
                "valores": fila_datos
            })
        
        wb.close()
        
        # Guardar
        mostrar_progreso_consola("Guardando JSON", 95, "Finalizando...")
        
        # Generar nombre por defecto
        nombre_base = os.path.splitext(os.path.basename(archivo_excel))[0]
        nombre_predeterminado = f"{nombre_base}_MUESTRA_{filas}f_{columnas}c_{time.strftime('%Y%m%d_%H%M%S')}.json"
        
        # Si no se proporcionó destino, preguntar al usuario
        if not destino_guardado:
            destino_guardado = seleccionar_destino_guardado(nombre_predeterminado)
            
            if not destino_guardado:  # Usuario canceló
                print("\n❌ Guardado cancelado por el usuario.")
                return None, muestra
        
        # Asegurar extensión .json
        if not destino_guardado.lower().endswith('.json'):
            destino_guardado += '.json'
        
        with open(destino_guardado, 'w', encoding='utf-8') as f:
            json.dump(muestra, f, ensure_ascii=False, indent=2)
        
        mostrar_progreso_consola("✅ COMPLETADO", 100, 
                               f"{len(muestra['datos'])} filas en {destino_guardado}")
        
        # Mostrar estadísticas
        print(f"\n📊 ESTADÍSTICAS FINALES:")
        print(f"   - Filas procesadas: {len(muestra['datos'])}")
        print(f"   - Columnas por fila: {columnas_a_procesar}")
        print(f"   - Primera etiqueta: {etiquetas[0]}")
        print(f"   - Última etiqueta: {etiquetas[-1]}")
        
        # Mostrar ejemplo de primera fila
        if muestra['datos']:
            print(f"\n📄 EJEMPLO (fila 1):")
            primera_fila = muestra['datos'][0]['valores']
            for key in list(primera_fila.keys())[:3]:
                print(f"   {key}: {primera_fila[key]}")
            print(f"   ... ({len(primera_fila)} campos totales)")
        
        return destino_guardado, muestra
        
    except Exception as e:
        raise Exception(f"Error: {str(e)}")

def preguntar_modo(total_filas, total_columnas):
    """Pregunta al usuario cómo quiere procesar el archivo"""
    root = tk.Tk()
    root.withdraw()
    
    # Ventana de selección de modo
    modo_window = tk.Toplevel(root)
    modo_window.title("Seleccionar Modo de Procesamiento")
    modo_window.geometry("550x400")
    
    # Centrar
    modo_window.update_idletasks()
    x = (modo_window.winfo_screenwidth() // 2) - 275
    y = (modo_window.winfo_screenheight() // 2) - 200
    modo_window.geometry(f'550x400+{x}+{y}')
    
    resultado = {"modo": None}
    
    tk.Label(modo_window, text="🔧 SELECCIONAR MODO DE PROCESAMIENTO", 
             font=("Arial", 14, "bold"), pady=15).pack()
    
    # Información del archivo
    info_frame = tk.Frame(modo_window, relief="groove", borderwidth=2, padx=10, pady=10)
    info_frame.pack(pady=10, padx=20, fill="x")
    
    tk.Label(info_frame, text=f"📁 Archivo con:", font=("Arial", 10, "bold")).pack(anchor='w')
    tk.Label(info_frame, text=f"   • {total_filas:,} filas totales".replace(",", "."), 
             font=("Arial", 9)).pack(anchor='w')
    tk.Label(info_frame, text=f"   • {total_columnas} columnas totales", 
             font=("Arial", 9)).pack(anchor='w')
    
    # Opciones
    opciones_frame = tk.Frame(modo_window, pady=15)
    opciones_frame.pack()
    
    def seleccionar_modo(modo):
        resultado["modo"] = modo
        modo_window.destroy()
    
    # Botón para Muestra Rápida (10,000 filas, 22 columnas específicas)
    btn_muestra = tk.Button(opciones_frame, text="⚡ MUESTRA RÁPIDA (10,000 filas)", 
                          command=lambda: seleccionar_modo("muestra_configurable"),
                          width=35, height=3, bg="#2196F3", fg="white",
                          font=("Arial", 11, "bold"))
    btn_muestra.pack(pady=10)
    
    tk.Label(opciones_frame, text="Hasta 10,000 filas con 22 etiquetas específicas", 
             font=("Arial", 9), fg="gray").pack()
    tk.Label(opciones_frame, text="CVE, NOMBRE, PATERNO, MATERNO, FECNAC, SEXO, ...", 
             font=("Arial", 8), fg="blue").pack()
    
    # Botón para Versión Completa
    btn_completo = tk.Button(opciones_frame, text="📈 VERSIÓN COMPLETA", 
                           command=lambda: seleccionar_modo("completo"),
                           width=35, height=3, bg="#4CAF50", fg="white",
                           font=("Arial", 11, "bold"))
    btn_completo.pack(pady=10)
    
    tk.Label(opciones_frame, text="Procesa TODAS las filas y columnas", 
             font=("Arial", 9), fg="gray").pack()
    
    # Botón Cancelar
    def cancelar():
        resultado["modo"] = None
        modo_window.destroy()
    
    tk.Button(modo_window, text="❌ Cancelar", command=cancelar,
              width=15).pack(pady=20)
    
    # Esperar a que se cierre la ventana
    modo_window.wait_window(modo_window)
    
    return resultado["modo"]

def mostrar_resultado(archivo_json, datos):
    """Muestra el resultado"""
    if not archivo_json:  # Usuario canceló el guardado
        return
    
    root = tk.Tk()
    root.title("✅ Conversión Completada")
    root.geometry("500x400")
    
    # Centrar
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - 250
    y = (root.winfo_screenheight() // 2) - 200
    root.geometry(f'500x400+{x}+{y}')
    
    tk.Label(root, text="✅ CONVERSIÓN EXITOSA", 
             font=("Arial", 14, "bold"), pady=20).pack()
    
    nombre = os.path.basename(archivo_json)
    tamano_kb = os.path.getsize(archivo_json) / 1024
    ubicacion = os.path.dirname(archivo_json)
    
    tk.Label(root, text=f"📁 Archivo: {nombre}", 
             font=("Arial", 10), wraplength=450).pack(pady=5)
    tk.Label(root, text=f"📊 Tamaño: {tamano_kb:.1f} KB", 
             font=("Arial", 10)).pack(pady=5)
    
    # Contar datos procesados
    if "hojas" in datos:  # Versión completa
        total_celdas = sum(h["celdas_procesadas"] for h in datos["hojas"])
        total_filas = sum(len(h["datos"]) for h in datos["hojas"])
        tk.Label(root, text=f"🔢 {total_filas:,} filas, {total_celdas:,} celdas".replace(",", "."), 
                 font=("Arial", 10, "bold")).pack(pady=10)
    elif "datos" in datos:  # Muestra rápida
        filas_proc = len(datos["datos"])
        columnas_proc = datos["configuracion"]["columnas_procesadas"]
        tk.Label(root, text=f"📈 {filas_proc:,} filas × {columnas_proc} columnas".replace(",", "."), 
                 font=("Arial", 10, "bold")).pack(pady=10)
        tk.Label(root, text=f"🏷️  Etiquetas: {', '.join(datos['etiquetas'][:4])}...", 
                 font=("Arial", 9)).pack(pady=5)
    
    tk.Label(root, text=f"📂 Ubicación: {ubicacion}", 
             font=("Arial", 9), fg="gray", wraplength=450).pack(pady=5)
    
    # Botones
    frame = tk.Frame(root, pady=25)
    frame.pack()
    
    def abrir_carpeta():
        os.startfile(os.path.dirname(archivo_json))
    
    def abrir_archivo():
        os.startfile(archivo_json)
    
    tk.Button(frame, text="📂 Abrir Carpeta", command=abrir_carpeta,
              width=18, height=2).pack(pady=5)
    tk.Button(frame, text="📄 Abrir JSON", command=abrir_archivo,
              width=18, height=2).pack(pady=5)
    tk.Button(frame, text="❌ Cerrar", command=root.destroy,
              width=18, height=2).pack(pady=5)
    
    root.mainloop()

def main():
    """Función principal"""
    print("=" * 60)
    print("CONVERSOR EXCEL A JSON - VERSIÓN MEJORADA")
    print("=" * 60)
    
    # Seleccionar archivo
    print("\n📂 Selecciona un archivo Excel...")
    root = tk.Tk()
    root.withdraw()
    
    archivo = filedialog.askopenfilename(
        title="Selecciona archivo Excel",
        filetypes=[("Excel files", "*.xlsx *.xls *.xlsm"), ("All files", "*.*")]
    )
    
    if not archivo:
        print("Operación cancelada.")
        return
    
    # Mostrar información del archivo
    tamano_mb = os.path.getsize(archivo) / (1024 * 1024)
    print(f"\n📁 Archivo seleccionado: {os.path.basename(archivo)}")
    print(f"📏 Tamaño: {tamano_mb:.2f} MB")
    
    try:
        # Obtener información básica del archivo
        wb = load_workbook(archivo, read_only=True, data_only=True)
        ws = wb.active
        total_filas = ws.max_row
        total_columnas = ws.max_column
        wb.close()
        
        print(f"📊 Contenido: {total_filas:,} filas × {total_columnas} columnas".replace(",", "."))
        print("-" * 60)
        
        # Preguntar modo de procesamiento
        modo = preguntar_modo(total_filas, total_columnas)
        
        if modo is None:
            print("❌ Operación cancelada por el usuario.")
            return
        
        try:
            if modo == "muestra_configurable":
                # Preguntar configuración específica (filas solamente, columnas fijas en 22)
                filas, columnas = preguntar_numero_filas_columnas(total_filas, total_columnas)
                
                if filas is None or columnas is None:
                    print("❌ Configuración cancelada.")
                    return
                
                print(f"\n⚙️ Configuración seleccionada:")
                print(f"   - Filas: {filas:,}".replace(",", "."))
                print(f"   - Columnas: {columnas} (etiquetas fijas)")
                
                archivo_json, datos = procesar_muestra_rapida(archivo, filas=filas, columnas=22)
                
            elif modo == "completo":
                print(f"\n⚙️ Configuración: Versión completa (todas las filas y columnas)")
                archivo_json, datos = procesar_excel_completo(archivo)
            
            if archivo_json:
                print(f"\n✅ Archivo JSON creado exitosamente!")
                print(f"📍 Ubicación: {archivo_json}")
                mostrar_resultado(archivo_json, datos)
            else:
                print("\n⚠️ Operación cancelada.")
                
        except Exception as e:
            print(f"\n❌ Error durante el procesamiento: {e}")
            messagebox.showerror("Error", f"No se pudo procesar el archivo:\n\n{str(e)}")
            
    except Exception as e:
        print(f"\n❌ Error al leer el archivo: {e}")
        messagebox.showerror("Error", f"No se pudo leer el archivo:\n\n{str(e)}")

if __name__ == "__main__":
    # Verificar dependencias
    try:
        import openpyxl
        import datetime
    except ImportError:
        print("Instalando dependencias...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        print("\n✅ Dependencias instaladas. Ejecuta el programa nuevamente.")
        input("Presiona Enter para salir...")
        sys.exit(0)
    
    # Ejecutar
    main()
